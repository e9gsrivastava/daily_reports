import csv
import gzip
import os
import zipfile
from datetime import date

from openpyxl import load_workbook

subfolder_headers = {
    "ercot_node_dam_lmp": {
        "headers": "DeliveryDate,HourEnding,SettlementPoint,SettlementPointPrice,DSTFlag",
        "Expected_file_count": 1,
        "expected_file_size": 140.8164,
    },
    "ercot_node_fmm_lmp": {
        "headers": "DeliveryDate,DeliveryHour,DeliveryInterval,SettlementPointName,SettlementPointType,SettlementPointPrice,DSTFlag",
        "Expected_file_count": 96,
        "expected_file_size": 541.0205,
    },
    "ercot_node_rtm_lmp": {
        "headers": "SCEDTimestamp,RepeatedHourFlag,SettlementPoint,LMP",
        "Expected_file_count": 288,
        "expected_file_size": 1589.1719,
    },
    "ercot_as_dam_regup_regdown_spin_nonspin_ecrs": {
        "headers": "DeliveryDate,HourEnding,AncillaryType,MCPC,DSTFlag",
        "Expected_file_count": 1,
        "expected_file_size": 0.8174,
    },
    "ercot_as_fmm_ordc": {
        "headers": "DeliveryDate,DeliveryHour,DeliveryInterval,RTRSVPOR,RTRSVPOFF,RTRDP,DSTFlag",
        "Expected_file_count": 96,
        "expected_file_size": 34.7266,
    },
    "ercot_busbar_dam_lmp": {
        "headers": "DeliveryDate,HourEnding,BusName,LMP,DSTFlag",
        "Expected_file_count": 1,
        "expected_file_size": 2024.7051,
    },
    "ercot_busbar_rtm_lmp": {
        "headers": "SCEDTimestamp,RepeatedHourFlag,ElectricalBus,LMP",
        "Expected_file_count": 288,
        "expected_file_size": 20963.4863,
    },
    "nyiso_hub_dam_lmp": {
        "headers": "Time Stamp,Name,PTID,LBMP ($/MWHr),Marginal Cost Losses ($/MWHr),Marginal Cost Congestion ($/MWHr)",
        "Expected_file_count": 1,
        "expected_file_size": 16.415,
    },
    "nyiso_hub_twha_lmp": {
        "headers": "Time Stamp,Name,PTID,LBMP ($/MWHr),Marginal Cost Losses ($/MWHr),Marginal Cost Congestion ($/MWHr)",
        "Expected_file_count": 1,
        "expected_file_size": 17.835,
    },
    "nyiso_hub_rtm_lmp": {
        "headers": "Time Stamp,Name,PTID,LBMP ($/MWHr),Marginal Cost Losses ($/MWHr),Marginal Cost Congestion ($/MWHr)",
        "Expected_file_count": 1,
        "expected_file_size": 225.4004,
    },
    "nyiso_node_dam_lmp": {
        "headers": "Time Stamp,Name,PTID,LBMP ($/MWHr),Marginal Cost Losses ($/MWHr),Marginal Cost Congestion ($/MWHr)",
        "Expected_file_count": 1,
        "expected_file_size": 1008.5391,
    },
    "nyiso_node_twha_lmp": {
        "headers": "Time Stamp,Name,PTID,LBMP ($/MWHr),Marginal Cost Losses ($/MWHr),Marginal Cost Congestion ($/MWHr)",
        "Expected_file_count": 1,
        "expected_file_size": 1077.0576,
    },
    "nyiso_node_rtm_lmp": {
        "headers": "Time Stamp,Name,PTID,LBMP ($/MWHr),Marginal Cost Losses ($/MWHr),Marginal Cost Congestion ($/MWHr)",
        "Expected_file_count": 1,
        "expected_file_size": 13334.9775,
    },
    "nyiso_as_dam_reg_spin_nonspin": {
        "headers": "Time Stamp,Time Zone,Name,PTID,10 Min Spinning Reserve ($/MWHr),10 Min Non-Synchronous Reserve ($/MWHr),30 Min Operating Reserve ($/MWHr),NYCA Regulation Capacity ($/MWHr)",
        "Expected_file_count": 1,
        "expected_file_size": 14.3105,
    },
    "nyiso_as_rtm_reg_spin_nonspin": {
        "headers": "Time Stamp,Time Zone,Name,PTID,10 Min Spinning Reserve ($/MWHr),10 Min Non-Synchronous Reserve ($/MWHr),30 Min Operating Reserve ($/MWHr),NYCA Regulation Capacity ($/MWHr),NYCA Regulation Movement ($/MW)",
        "Expected_file_count": 1,
        "expected_file_size": 194.5908,
    },
    "isone_node_dam_lmp": {
        "headers": "C,Day-Ahead Energy Market Hourly LMP Report",
        "Expected_file_count": 1,
        "expected_file_size": 2418.8555,
    },
    "isone_node_twha_lmp": {
        "headers": "C,Real-Time Energy Market Hourly LMP Report",
        "Expected_file_count": 1,
        "expected_file_size": 2358.6484,
    },
    "isone_node_rtm_lmp": {
        "headers": "C,Five-Minute Locational Marginal Prices",
        "Expected_file_count": 6,
        "expected_file_size": 13547.5488,
    },
    "isone_as_rtm_reg": {
        "headers": "C,Final Real-Time Five-Minute Regulation Clearing Prices",
        "Expected_file_count": 1,
        "expected_file_size": 10.0566,
    },
    "isone_as_twha_spin_nonspin": {
        "headers": "C,Final Hourly Reserve Zone Prices & Designations",
        "Expected_file_count": 1,
        "expected_file_size": 5.4619,
    },
    "isone_as_rtm_spin_nonspin": {
        "headers": "C,Five-Minute Reserve Zone Prices & Designation final version",
        "Expected_file_count": 1,
        "expected_file_size": 109.0957,
    },
    "caiso_node_dam_lmp": {
        "headers": "INTERVALSTARTTIME_GMT,INTERVALENDTIME_GMT,OPR_DT,OPR_HR,OPR_INTERVAL,NODE_ID_XML,NODE_ID,NODE,MARKET_RUN_ID,LMP_TYPE,XML_DATA_ITEM,PNODE_RESMRID,GRP_TYPE,POS,MW,GROUP",
        "Expected_file_count": 1,
        "expected_file_size": 13229.3486,
    },
    "caiso_node_rtm_lmp": {
        "headers": "INTERVALSTARTTIME_GMT,INTERVALENDTIME_GMT,OPR_DT,OPR_HR,NODE_ID_XML,NODE_ID,NODE,MARKET_RUN_ID,LMP_TYPE,XML_DATA_ITEM,PNODE_RESMRID,GRP_TYPE,POS,VALUE,OPR_INTERVAL,GROUP",
        "Expected_file_count": 24,
        "expected_file_size": 218985.5537,
    },
    "caiso_node_fmm_lmp": {
        "headers": "INTERVALSTARTTIME_GMT,INTERVALENDTIME_GMT,OPR_DT,OPR_HR,NODE_ID_XML,NODE_ID,NODE,MARKET_RUN_ID,LMP_TYPE,XML_DATA_ITEM,PNODE_RESMRID,GRP_TYPE,POS,PRC,OPR_INTERVAL,GROUP",
        "Expected_file_count": 24,
        "expected_file_size": 84019.3682,
    },
    "caiso_as_dam_regup_regdown_spin_nonspin": {
        "headers": "INTERVALSTARTTIME_GMT,INTERVALENDTIME_GMT,OPR_DT,OPR_HR,OPR_INTERVAL,OPR_TYPE,ANC_TYPE,ANC_REGION,MARKET_RUN_ID,XML_DATA_ITEM,MW,GROUP",
        "Expected_file_count": 1,
        "expected_file_size": 3.4531,
    },
    "caiso_as_fmm_regup_regdown_spin_nonspin": {
        "headers": "INTERVALSTARTTIME_GMT,INTERVALENDTIME_GMT,OPR_DT,OPR_HR,OPR_TYPE,ANC_TYPE,XML_DATA_ITEM,MARKET_RUN_ID,MW,ANC_REGION,OPR_INTERVAL,GROUP",
        "Expected_file_count": 24,
        "expected_file_size": 22.6699,
    },
    "miso_node_dam_lmp": {
        "headers": "Day Ahead Market ExPost LMPs",
        "Expected_file_count": 1,
        "expected_file_size": 990.4775,
    },
    "miso_node_rtm_lmp": {
        "headers": "Real-Time Ex-Ante 5 Minute LMP Report",
        "Expected_file_count": 1,
        "expected_file_size": 19354.2461,
    },
    "miso_node_twha_lmp": {
        "headers": "Realtime Market LMPs",
        "Expected_file_count": 1,
        "expected_file_size": 1148.1191,
    },
    "miso_as_dam_reg_spin_nonspin": {
        "headers": "Dayahead Market MCPs.",
        "Expected_file_count": 1,
        "expected_file_size": 608.2021,
    },
    "miso_as_rtm_reg_spin_nonspin": {
        "headers": "Real-Time Ex-Ante 5 Minute MCP Report",
        "Expected_file_count": 1,
        "expected_file_size": 61.9141,
    },
    "miso_as_twha_reg_spin_nonspin": {
        "headers": "Real Time Market MCPs.,Not settlement quality data.",
        "Expected_file_count": 1,
        "expected_file_size": 695.9531,
    },
    "miso_busbar_dam_lmp": {
        "headers": "Day-Ahead EPNode Price data",
        "Expected_file_count": 1,
        "expected_file_size": 1177.4326,
    },
    "miso_busbar_twha_lmp": {
        "headers": "Real-Time EPNode Price data",
        "Expected_file_count": 1,
        "expected_file_size": 1114.0225,
    },
    "pjm_node_dam_lmp": {
        "headers": "datetime_beginning_utc,datetime_beginning_ept,pnode_id,pnode_name,voltage,equipment,type,zone,system_energy_price_da,total_lmp_da,congestion_price_da,marginal_loss_price_da,row_is_current,version_nbr",
        "Expected_file_count": 1,
        "expected_file_size": 4501.0762,
    },
    "pjm_node_rtm_lmp": {
        "headers": "datetime_beginning_utc,datetime_beginning_ept,pnode_id,pnode_name,voltage,equipment,type,zone,system_energy_price_rt,total_lmp_rt,congestion_price_rt,marginal_loss_price_rt,row_is_current,version_nbr",
        "Expected_file_count": 4,
        "expected_file_size": 49768.8008,
    },
    "pjm_node_twha_lmp": {
        "headers": "datetime_beginning_utc,datetime_beginning_ept,pnode_id,pnode_name,voltage,equipment,type,zone,system_energy_price_rt,total_lmp_rt,congestion_price_rt,marginal_loss_price_rt,row_is_current,version_nbr",
        "Expected_file_count": 1,
        "expected_file_size": 4310.8613,
    },
    "pjm_as_dam_spin_nonspin": {
        "headers": "datetime_beginning_utc,datetime_beginning_ept,ancillary_service,unit,value,row_is_current,version_nbr",
        "Expected_file_count": 1,
        "expected_file_size": 0.751,
    },
    "pjm_as_rtm_reg_spin_nonspin": {
        "headers": "datetime_beginning_utc,datetime_beginning_ept,ancillary_service,unit,value,version_nbr,row_is_current",
        "Expected_file_count": 1,
        "expected_file_size": 10.1123,
    },
    "pjm_as_twha_reg_spin_nonspin": {
        "headers": "datetime_beginning_utc,datetime_beginning_ept,ancillary_service,unit,value,row_is_current,version_nbr",
        "Expected_file_count": 1,
        "expected_file_size": 1.3838,
    },
}

today_date = date.today()
root_dir = "/home/fox/Desktop/july_onwards_daily_reports/daily_data/20240530"
output_csv = f"/home/fox/Desktop/july_onwards_daily_reports/file_count/{today_date}_daily_report.csv"


def read_file_headers(file_path):
    try:
        with open(file_path, "r", encoding="utf-8") as file:
            first_line = file.readline().strip().replace('"', "")
        return first_line
    except UnicodeDecodeError:
        return None


def read_headers_from_file(file_path):
    _, file_extension = os.path.splitext(file_path)
    if file_extension == ".csv":
        return read_file_headers(file_path)
    elif file_extension == ".xlsx":
        try:
            wb = load_workbook(filename=file_path, read_only=True)
            ws = wb.active
            header_row = next(
                ws.iter_rows(min_row=1, max_row=1, values_only=True)
            )
            return ",".join(
                str(cell).replace('"', "") for cell in header_row if cell
            )
        except Exception as e:
            print(f"Error reading headers from {file_path}: {e}")
            return None
    elif file_extension == ".zip":
        try:
            with zipfile.ZipFile(file_path, "r") as zip_ref:
                for file in zip_ref.namelist():
                    if file.endswith(".csv"):
                        csv_data = zip_ref.read(file).decode("utf-8")
                        first_line = (
                            csv_data.splitlines()[0].strip().replace('"', "")
                        )
                        return first_line
            return None
        except Exception as e:
            print(f"Error reading headers from {file_path}: {e}")
            return None
    elif file_extension == ".gz":
        try:
            with gzip.open(file_path, "rb") as gz_file:
                csv_data = gz_file.read().decode("utf-8")
                first_line = csv_data.splitlines()[0].strip().replace('"', "")
            return first_line
        except Exception as e:
            print(f"Error reading headers from {file_path}: {e}")
            return None
    else:
        print(f"Unsupported file extension for {file_path}")
        return None


def compare_headers(subfolder_path, expected_headers):
    for file_name in os.listdir(subfolder_path):
        file_path = os.path.join(subfolder_path, file_name)
        actual_headers = read_headers_from_file(file_path)

        print("************************************************************")
        print(actual_headers)
        print(expected_headers)

        print("************************************************************")
        if actual_headers is None:
            print(f"Could not read headers from {file_path}")
            return False
        if actual_headers != expected_headers:
            return False
    return True


def get_directory_size(directory):
    total_size = 0
    for dirpath, _, filenames in os.walk(directory):
        for filename in filenames:
            filepath = os.path.join(dirpath, filename)
            total_size += os.path.getsize(filepath)
    return round(total_size / 1024, 4)


results = []

for subfolder, expected_info in subfolder_headers.items():
    subfolder_path = os.path.join(root_dir, subfolder)
    if os.path.exists(subfolder_path):
        headers_match = compare_headers(
            subfolder_path, expected_info["headers"]
        )
        if headers_match:
            results.append((subfolder, "match"))
        else:
            results.append((subfolder, "mismatch"))

data = {}


for subfolder, expected_info in subfolder_headers.items():
    subfolder_path = os.path.join(root_dir, subfolder)
    if os.path.exists(subfolder_path):
        size_kb = get_directory_size(subfolder_path)
        file_count = len(os.listdir(subfolder_path))
        data[subfolder] = {
            "size_kb": size_kb,
            "file_count": file_count,
            "Expected_file_count": expected_info["Expected_file_count"],
            "expected_file_size": expected_info["expected_file_size"],
        }
    else:
        print(
            f"Subfolder '{subfolder}' does not exist or could not be accessed."
        )
        data[subfolder] = {
            "size_kb": 0,
            "file_count": 0,
            "Expected_file_count": expected_info["Expected_file_count"],
            "expected_file_size": expected_info["expected_file_size"],
        }

with open(output_csv, "w", newline="") as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(
        [
            "report_id",
            "folder_size_kb",
            "file_count",
            "Expected_file_count",
            "expected_folder_size",
            "headers_match_or_not",
        ]
    )
    for subfolder in subfolder_headers.keys():
        headers_match = next(
            (status for (folder, status) in results if folder == subfolder),
            "not checked",
        )
        writer.writerow(
            [
                subfolder,
                data[subfolder]["size_kb"],
                data[subfolder]["file_count"],
                data[subfolder]["Expected_file_count"],
                data[subfolder]["expected_file_size"],
                headers_match,
            ]
        )

print(f"Combined CSV file generated: {output_csv}")
